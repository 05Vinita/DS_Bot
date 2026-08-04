"""
utils.py
--------
Helper functions that don't belong specifically to the Google Maps
scraper or the email extractor:
  - removing duplicate business records
  - saving the final dataset to Excel + CSV
"""

import os
from datetime import datetime

import pandas as pd

import config


def remove_duplicates(records):
    """
    Removes duplicate business entries.

    Two records are duplicates if they share the same Company Name AND
    Phone Number. Name alone is not reliable (a chain has the same name in
    several places) and phone alone is not either (some listings have none).

    Note the gap this leaves: two records with the same name and NO phone
    number both key on ("name", "") and collapse into one, even if they are
    genuinely different branches. Address is used as a tiebreaker in that
    case so real branches survive.

    Args:
        records (list[dict])

    Returns:
        list[dict]: de-duplicated list, order preserved
    """
    seen = set()
    unique_records = []

    for record in records:
        # str() first: .get(key, "") still returns None when the key exists
        # with a None value, and .strip() on None raises AttributeError
        key = (
            str(record.get("Company Name") or "").strip().lower(),
            str(record.get("Phone Number") or "").strip(),
        )
        # With no phone to tell them apart, fall back to the address so two
        # real branches of the same chain are not merged into one
        if not key[1]:
            key = key + (str(record.get("Address") or "").strip().lower(),)

        if key in seen:
            continue
        seen.add(key)
        unique_records.append(record)

    removed_count = len(records) - len(unique_records)
    if removed_count:
        print(f"Removed {removed_count} duplicate record(s).")

    return unique_records


def _merge_into_combined(df, combined_path):
    """
    Adds this run's rows to the combined file for this keyword and city.

    Reads what is already there, appends, and drops duplicates on Company
    Name + Phone Number -- the same rule remove_duplicates() uses, so a
    business that appears in two runs is stored once.

    Returns the merged DataFrame, or None if it could not be written.
    """
    try:
        if os.path.exists(combined_path):
            # dtype=str on the way in: without it pandas reads "+919812345678"
            # back as a float and the leading "+" is lost, so every merge
            # would quietly degrade the phone columns a little further.
            existing = pd.read_excel(combined_path, dtype=str).fillna("")
            merged = pd.concat([existing, df], ignore_index=True)
        else:
            merged = df.copy()

        for column in ("Phone Number", "WhatsApp", "Pin Code"):
            if column in merged.columns:
                merged[column] = merged[column].fillna("").astype(str)
                # ".0" is what an int looks like after a float round-trip
                merged[column] = merged[column].str.replace(
                    r"\.0$", "", regex=True
                )

        # Normalise before comparing: pandas reads a blank cell back as NaN,
        # so a row saved with "" and the same row reloaded would not match
        # and the duplicate would survive.
        for column in ("Company Name", "Phone Number"):
            if column in merged.columns:
                merged[column] = (
                    merged[column].fillna("").astype(str).str.strip()
                )

        if "Company Name" in merged.columns:
            merged["_dedup_name"] = merged["Company Name"].str.lower()
        if "Phone Number" in merged.columns:
            merged["_dedup_phone"] = merged["Phone Number"].str.replace(
                r"\D", "", regex=True
            )

        subset = [c for c in ("_dedup_name", "_dedup_phone") if c in merged.columns]
        if subset:
            merged = merged.drop_duplicates(subset=subset, keep="first")
        merged = merged.drop(columns=[c for c in subset], errors="ignore")

        merged.to_excel(combined_path, index=False)
        _format_text_columns(combined_path)
        return merged
    except Exception as err:
        print(f"  (could not update the combined file: {err})")
        return None


def _format_text_columns(xlsx_path):
    """Keeps phone-like columns as text so Excel does not mangle them."""
    try:
        import openpyxl
        book = openpyxl.load_workbook(xlsx_path)
        sheet = book.active
        headers = [cell.value for cell in sheet[1]]
        for name in ("Phone Number", "WhatsApp", "Pin Code"):
            if name in headers:
                index = headers.index(name) + 1
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=index).number_format = "@"
        book.save(xlsx_path)
    except Exception as err:
        print(f"  (could not set text format on phone columns: {err})")


def save_results(records, keyword, city):
    """
    Saves the final list of business records to both .xlsx and .csv,
    using the naming convention: keyword_city.xlsx / keyword_city.csv

    Args:
        records (list[dict])
        keyword (str)
        city (str)

    Returns:
        tuple[str, str]: (path_to_xlsx, path_to_csv)
    """
    if not records:
        print("No records to save.")
        return None, None

    df = pd.DataFrame(records)

    # Force phone-like columns to stay text.
    #
    # Without this, pandas sees a column of digits and stores it as a number:
    # the leading "+" disappears and Excel shows 9.19812e+11 instead of
    # +919812345678. The value is then useless for copying into WhatsApp,
    # which is the whole point of the column.
    for column in ("Phone Number", "WhatsApp", "Pin Code"):
        if column in df.columns:
            df[column] = df[column].fillna("").astype(str)

    # Keep a consistent, readable column order
    preferred_order = [
        "Company Name",
        "Category",
        "Phone Number",
        "WhatsApp",
        "Email",
        "Address",
        "Website",
        "Rating",
        "City",
        "State",
        "Pin Code",
        "Instagram",
        "Facebook",
        "LinkedIn",
        "YouTube",
    ]
    existing_cols = [c for c in preferred_order if c in df.columns]
    other_cols = [c for c in df.columns if c not in existing_cols]
    df = df[existing_cols + other_cols]

    # Build the filename, e.g. "civil_jaipur"
    safe_keyword = keyword.strip().lower().replace(" ", "_")
    safe_city = city.strip().lower().replace(" ", "_")
    base_filename = f"{safe_keyword}_{safe_city}"

    # A second run for the same keyword and city holds only the NEW leads --
    # the earlier ones were skipped on purpose. Writing to the same filename
    # would therefore not "update" the file, it would replace 180 leads with
    # whatever this run happened to add. So each run gets its own dated file,
    # and a combined file is kept alongside it.
    # Seconds included on purpose: two runs started inside the same minute
    # would otherwise collide and reintroduce the very overwrite this avoids.
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    run_filename = f"{base_filename}_{stamp}"

    # Even to the second, two runs can collide -- a fast re-run, or a script
    # calling this twice. Losing a run's output to a name clash is exactly
    # the bug this whole scheme exists to prevent, so check and step aside.
    suffix = 2
    while os.path.exists(os.path.join(config.OUTPUT_DIR, f"{run_filename}.xlsx")):
        run_filename = f"{base_filename}_{stamp}_{suffix}"
        suffix += 1

    # gmaps_scraper creates this too, but save_results can be called on its
    # own, and pandas raises rather than creating a missing folder
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    xlsx_path = os.path.join(config.OUTPUT_DIR, f"{run_filename}.xlsx")
    csv_path = os.path.join(config.OUTPUT_DIR, f"{run_filename}.csv")

    df.to_excel(xlsx_path, index=False)
    df.to_csv(csv_path, index=False)

    # The combined file: every run for this keyword and city, deduplicated.
    # This is the one to actually work from.
    combined_path = os.path.join(config.OUTPUT_DIR, f"{base_filename}_ALL.xlsx")
    combined = _merge_into_combined(df, combined_path)

    # Excel guesses types from cell content when a file is opened, so a
    # column of digits can still be reformatted on the way in. Marking the
    # cells as text explicitly keeps "+91..." intact.
    _format_text_columns(xlsx_path)

    print(f"Saved {len(df)} new records from this run to:")
    print(f"  - {xlsx_path}")
    print(f"  - {csv_path}")
    if combined is not None:
        print(f"\nAll {len(combined)} records for '{keyword}' in {city} "
              "(this run plus earlier ones):")
        print(f"  - {combined_path}")

    return xlsx_path, csv_path